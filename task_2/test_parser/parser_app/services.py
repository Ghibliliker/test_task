import openpyxl
import asyncio
import json
import aiohttp
from asgiref.sync import async_to_sync

from parser_app.schemas import ItemBase


class XlsxDecode:

    @staticmethod
    def decode(data):
        ans = []
        wookbook = openpyxl.load_workbook(data)
        worksheet = wookbook.active
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                ans.append(str(col[i].value))
        return ans


class ParseService:

        result = []

        @staticmethod
        async def start_parse(list_id):

            async def fetch_data(url, session):
                async with session.get(url) as response:
                    ans = await response.read()
                    data = json.loads(ans.decode('utf-8'))
                    item = ItemBase(
                        article=data['nm_id'],
                        brand=data['imt_name'],
                        title=data['selling']['brand_name']
                    )
                    ParseService.result.append(item)

            async def bound_fetch(sem, url, session):
                async with sem:
                    print("doing request for " + url)
                    await fetch_data(url, session)

            async def run():
                tasks = []
                sem = asyncio.Semaphore(100)
                async with aiohttp.ClientSession() as session:
                    for id in list_id:
                        task = asyncio.ensure_future(
                            bound_fetch(
                                sem,
                                f'https://basket-05.wb.ru/vol{id[:3]}/part{id[:5]}/{id}/info/ru/card.json',
                                session
                            )
                        )
                        tasks.append(task)
                    await asyncio.gather(*tasks)
            await run()

        @staticmethod
        def show_data(list_id):
            async_to_sync(ParseService.start_parse)(list_id)
            return ParseService.result
